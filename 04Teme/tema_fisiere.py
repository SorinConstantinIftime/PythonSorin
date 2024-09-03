#definire in formal excel
"""
in terminal se executa comanda: pip install openpyxl
"""
import openpyxl
import os
from datetime import datetime
import pandas as pd

#2 Definire nume fisiere Excel pentru categorii si task-uri
categorii_file = 'categorii.xlxs'
taskuri_file = 'taskuri.xlxs'

#3 verificare existenta fisiere si/sau creare, salvarea si inchiderea fisisrelor
if not os.path.exists(categorii_file) or not os.path.exists(taskuri_file):
    workbook = openpyxl.workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="lista categorii")
    workbook.close()





    sheet.cell(row=1,column=2,value="Data linita")
    sheet.cell(row=1,column=3,value="Persoana responsabila")
    sheet.cell(row=1,column=4, value="Categorie")
    workbook.save(taskuri_file)
    workbook.close()


#4. Functie pentru a adauga o noua categorie
def adauga_categorie():
    categorii_workbook = openpyxl.load_workbook(categorii_file)
